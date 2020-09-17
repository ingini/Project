import pyairmore
from openpyxl import load_workbook
from ipaddress import IPv4Address  # for your IP address
from pyairmore.request import AirmoreSession  # to create an AirmoreSession
from pyairmore.services.messaging import MessagingService  # to send messages

# Airmore 활성화
#ipAddress = input("핸드폰과 연결할 내부아이피를 입력하세요. ")

if not ipAddress:
    #ip = IPv4Address("192.168.100.210")

else:
    ip = IPv4Address(ipAddress)

session = AirmoreSession(ip, 2333)
service = MessagingService(session)

# 보낼 문자내용
filepathMessage = open("C:/Users/quantec/Desktop/kiwoom/groupSendSMS/message.txt", 'rt', encoding='UTF8')
message = filepathMessage.read()
filepathMessage.close()

# 주소록 파일
filepathContact = "C:/Users/quantec/Desktop/kiwoom/groupSendSMS/contact.xlsx"

# 연락처 컬럼
columnName = "A"
columnPhone = "B"

workbook = load_workbook(filename=filepathContact, read_only=True)
worksheet = workbook.worksheets[0]  # 첫번째 시트를 찾음

# 행과 열의 갯수를 찾음
row_count = worksheet.max_row
column_count = worksheet.max_column

phoneNumbers = {}

for i in range(row_count):
    cellPhone = "{}{}".format(columnPhone, i + 1)
    cellName = "{}{}".format(columnName, i + 1)
    name = worksheet[cellName].value
    number = worksheet[cellPhone].value
    if number != "" or number is not None:
        phoneNumbers[name] = str(number)

for nm, pn in phoneNumbers.items():
    sendMessage = "{name}님 안녕하세요. {message}".format(name=nm, message=message)
    print(nm)
    print(pn)
    print(sendMessage)
    print("")

print("총 {count}명에게 문자메시지 발송 가능합니다.\n".format(count=len(phoneNumbers)))
'''
order = input("위의 연락처로 문자를 발송하시겠습니까? Y or N ")

if order == "Y" or order == "YES" or order == "y" or order == "yes" or order == "Yes":
    for nm, pn in phoneNumbers.items():
        sendMessage = "{name}님 안녕하세요. {message}".format(name=nm, message=message)
        service.send_message(pn, sendMessage)
        print("{name}님의 {number}로 문자를 발송하였습니다.".format(name=nm, number=pn))
else:
    print("발송이 취소되었습니다.")
'''
for nm, pn in phoneNumbers.items():
    sendMessage = "{name}님 안녕하세요. {message}".format(name=nm, message=message)
    service.send_message(pn, sendMessage)
    print("{name}님의 {number}로 문자를 발송하였습니다.".format(name=nm, number=pn))
#input("아무 키나 누르면 종료합니다. ")